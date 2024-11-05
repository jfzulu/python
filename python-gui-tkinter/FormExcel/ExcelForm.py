import os.path
import tkinter as tk
from tkinter import messagebox
from openpyxl import Workbook, load_workbook
import re
import os

name_file="datos.xlsx"
#Verify is file exist
if os.path.exists(name_file):
    wb=load_workbook(name_file)
    ws=wb.active

else:
#Create Excel Book
    wb = Workbook()
    ws = wb.active
    ws.append(["Name", "Age", "Email", "Phone", "Address"])

def save_data():
    name= entry_name.get()
    age=entry_age.get()
    email=entry_email.get()
    phone=entry_phone.get()
    address=entry_address.get()

    if not name or not age or not email or not phone or not address:
        messagebox.showwarning("Warning", "All fields are required")
        return

    try:
        age = int(age)
        phone=int(phone)
    except ValueError:
        messagebox.showwarning("Warning", "Age and phone number must be numeric")

    if not re.match(r"[^@]+@[^@]+\.[^@]+", email):
        messagebox.showwarning("Warning", "Invalid Email")
        return

    ws.append([name, age, email, phone, address])
    wb.save(name_file)
    messagebox.showinfo("Information", "Data saved successfully")

    entry_name.delete(0, tk.END)
    entry_age.delete(0, tk.END)
    entry_phone.delete(0, tk.END)
    entry_email.delete(0, tk.END)
    entry_address.delete(0, tk.END)




root = tk.Tk()
root.title("Form Data Input")
root.configure(bg="#123452")
label_style = {"bg": '#123452', "fg": "white"}
entry_style = {"bg": '#234155', "fg":'white'}

#Get Name Label
label_name = tk.Label(root , text="Name", **label_style)
label_name.grid(row=0, column=0, padx=10, pady=5)
entry_name = tk.Entry(root, **entry_style)
entry_name.grid(row=0, column=1, padx=10, pady=5)

#Get Age Label
label_age = tk.Label(root, text="Age", **label_style)
label_age.grid(row=1, column=0, padx=10, pady=5)
entry_age = tk.Entry(root, **entry_style)
entry_age.grid(row=1, column=1, padx=10, pady=5)

#Get Email Label
label_email = tk.Label(root, text="Email", **label_style)
label_email.grid(row=2, column=0, padx=10, pady=5)
entry_email = tk.Entry(root, **entry_style)
entry_email.grid(row=2, column=1, padx=10, pady=5)

#Get Phone Label
label_phone = tk.Label(root, text="Phone", **label_style)
label_phone.grid(row=3, column=0, padx=10, pady=5)
entry_phone = tk.Entry(root, **entry_style)
entry_phone.grid(row=3, column=1, padx=10, pady=5)

#Get Address Label
label_address = tk.Label(root, text="Address", **label_style)
label_address.grid(row=4, column=0, padx=10, pady=5)
entry_address = tk.Entry(root, **entry_style)
entry_address.grid(row=4, column=1, padx=10, pady=5)

button_save=tk.Button(root, text="Save", command=save_data, bg='#6D8299', fg='white')
button_save.grid(row=5, column=0, columnspan=2, padx=10, pady=10)



root.mainloop()